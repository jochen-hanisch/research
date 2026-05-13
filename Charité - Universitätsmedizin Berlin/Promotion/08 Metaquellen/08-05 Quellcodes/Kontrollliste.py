import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

# Stimuli
stimuli = [
    "F2-S2", "F3-S3", "F5-S1", "F6-S1", "F8-S3",
    "F9-S3", "F10-S3", "F11-S3", "F12-S3-2",
    "F13-S3", "F14-S2"
]

# Kurskategorien
kurskategorien = ["21-NFS-09", "22-NFS-09", "23-NFS-09"]

# Geschlechterkategorien
geschlecht = ["Weiblich", "Männlich"]

# Ansichten
ansichten = ["Heatmap", "View-Map", "Fog-View"]

# Workbook und Worksheet erstellen
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Kontrollliste"

# Überschriften
headers = ["Stimulus", "Ansicht", "Kategorie", "Heruntergeladen", "Dateibenennung"]
ws.append(headers)

# Daten einfügen für Kurskategorien und Geschlecht
for stimulus in stimuli:
    for ansicht in ansichten:
        for kurskategorie in kurskategorien:
            for sex in geschlecht:
                dateiname = f"{stimulus}_{ansicht}_{kurskategorie}_{sex}".replace(" ", "-")
                ws.append([stimulus, ansicht, f"{kurskategorie} - {sex}", "", dateiname])
            dateiname = f"{stimulus}_{ansicht}_{kurskategorie}_Gesamt".replace(" ", "-")
            ws.append([stimulus, ansicht, f"{kurskategorie} - Gesamt", "", dateiname])
        dateiname = f"{stimulus}_{ansicht}_Gesamt".replace(" ", "-")
        ws.append([stimulus, ansicht, "Gesamt", "", dateiname])

# Checkboxen hinzufügen
dv = DataValidation(type="list", formula1='"Ja,Nein"', showDropDown=True)
ws.add_data_validation(dv)

# Bereich für Checkboxen definieren und anwenden
for row in range(2, ws.max_row + 1):
    cell = f"D{row}"
    dv.add(ws[cell])

# Workbook speichern
wb.save("Kontrollliste_EyeTracking.xlsx")

print("Die Kontrollliste mit Checkboxes wurde erfolgreich erstellt und in 'Kontrollliste_EyeTracking.xlsx' gespeichert.")